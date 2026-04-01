"""
Vistage - Grupos en Riesgo | App Web
Conecta a Zoho Creator y muestra el reporte con Risk Score
"""

import streamlit as st
import requests
import pandas as pd
from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuración de página ───────────────────────────────────────
st.set_page_config(
    page_title="Vistage · Grupos en Riesgo",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── Estilos ───────────────────────────────────────────────────────
st.markdown("""
<style>
    .main { background-color: #f8f9fa; }
    .metric-card {
        background: white;
        border-radius: 10px;
        padding: 20px;
        box-shadow: 0 2px 8px rgba(0,0,0,0.08);
        text-align: center;
    }
    .nivel-critico  { background:#C00000; color:white; padding:4px 12px; border-radius:20px; font-weight:bold; font-size:13px; }
    .nivel-alto     { background:#FF8C00; color:white; padding:4px 12px; border-radius:20px; font-weight:bold; font-size:13px; }
    .nivel-moderado { background:#FFD700; color:black; padding:4px 12px; border-radius:20px; font-weight:bold; font-size:13px; }
    .nivel-bajo     { background:#70AD47; color:white; padding:4px 12px; border-radius:20px; font-weight:bold; font-size:13px; }
    .header-box {
        background: linear-gradient(135deg, #1F4E79 0%, #2E75B6 100%);
        color: white;
        padding: 24px 32px;
        border-radius: 12px;
        margin-bottom: 24px;
    }
</style>
""", unsafe_allow_html=True)

# ── Credenciales desde Streamlit Secrets ─────────────────────────
def get_secret(key, default=""):
    try:
        return st.secrets[key]
    except:
        return default

CLIENT_ID     = get_secret("ZOHO_CLIENT_ID")
CLIENT_SECRET = get_secret("ZOHO_CLIENT_SECRET")
REFRESH_TOKEN = get_secret("ZOHO_REFRESH_TOKEN")
ZOHO_OWNER    = get_secret("ZOHO_OWNER", "ygomez_vistage")
APP_NAME      = get_secret("ZOHO_APP_NAME", "vistage-per")

REPORTES = {
    "grupos":   get_secret("REPORTE_GRUPOS", ""),
    "miembros": get_secret("REPORTE_MIEMBROS", ""),
    "bajas":    get_secret("REPORTE_BAJAS", ""),
    "chairs":   get_secret("REPORTE_CHAIRS", ""),
}

# ── Zoho API ──────────────────────────────────────────────────────
def get_access_token():
    r = requests.post("https://accounts.zoho.com/oauth/v2/token", data={
        "refresh_token": REFRESH_TOKEN,
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type": "refresh_token"
    })
    data = r.json()
    token = data.get("access_token")
    api_domain = data.get("api_domain", "https://www.zohoapis.com")
    return token, api_domain

def fetch_report(report_name, token):
    if not report_name:
        return []
    headers = {"Authorization": f"Zoho-oauthtoken {token}"}
    url = f"https://creator.zohoapis.com/api/v2/{ZOHO_OWNER}/{APP_NAME}/report/{report_name}"
    all_records, offset = [], 0
    while True:
        r = requests.get(url, headers=headers, params={"max_records": 200, "start_index": offset})
        data = r.json()
        records = data.get("data", [])
        all_records.extend(records)
        if len(records) < 200:
            break
        offset += 200
    return all_records

def discover_reports(token):
    headers = {"Authorization": f"Zoho-oauthtoken {token}"}
    url = f"https://creator.zohoapis.com/api/v2/{ZOHO_OWNER}/{APP_NAME}/report"
    r = requests.get(url, headers=headers)
    data = r.json()
    return [rep.get("link_name", "") for rep in data.get("reports", [])]

# ── Risk Score ────────────────────────────────────────────────────
def gv(rec, *keys):
    for k in keys:
        v = rec.get(k)
        if v not in (None, "", "null"):
            return v
    return None

def to_num(v, default=0):
    try:
        return float(str(v).replace(",", "."))
    except:
        return default

def calcular_puntaje(rec):
    antig    = int(to_num(gv(rec, "Antiguedad","Antigüedad","antiguedad"), 0))
    retencion= to_num(gv(rec, "Retencion","Retención","retencion"), 100)
    bajas_em = to_num(gv(rec, "Bajas_ene_mar","Bajas_ene-mar","bajas_ene_mar"), 0)
    bajas_od = to_num(gv(rec, "Bajas_oct_dic","Bajas_oct-dic","bajas_oct_dic"), 0)

    b1 = 20 if antig == 0 else (10 if antig == 1 else 0)
    if antig <= 1:
        b2 = 0 if bajas_em==0 else (15 if bajas_em==1 else (25 if bajas_em==2 else 35))
    else:
        b2 = 0 if bajas_em==0 else (5 if bajas_em==1 else (10 if bajas_em==2 else 15))
    b3 = 0 if retencion>=90 else (5 if retencion>=75 else (10 if retencion>=60 else (15 if retencion>=40 else 20)))
    b4 = 15 if bajas_em>bajas_od else (5 if bajas_em==bajas_od and bajas_em>0 else 0)
    b5 = 0
    total = b1+b2+b3+b4+b5
    nivel = "Crítico" if total>55 else ("Alto" if total>35 else ("Moderado" if total>15 else "Bajo"))
    return b1, b2, b3, b4, b5, total, nivel

NIVEL_EMOJI = {"Crítico":"🔴","Alto":"🟠","Moderado":"🟡","Bajo":"🟢"}
NIVEL_ORDER = {"Crítico":4,"Alto":3,"Moderado":2,"Bajo":1}

def build_dataframe(registros):
    rows = []
    for rec in registros:
        b1,b2,b3,b4,b5,total,nivel = calcular_puntaje(rec)
        rows.append({
            "Grupo":       gv(rec,"Nombre","nombre","Name") or "",
            "Programa":    gv(rec,"Programa_Vistage","Programa","programa") or "",
            "Chair":       gv(rec,"Chair","chair") or "",
            "Antigüedad":  int(to_num(gv(rec,"Antiguedad","Antigüedad"),0)),
            "Miembros":    int(to_num(gv(rec,"Miembros_pagantes","MiembrosPagantes"),0)),
            "Total":       int(to_num(gv(rec,"Total_Miembros","TotalMiembros"),0)),
            "Retención %": round(to_num(gv(rec,"Retencion","Retención"),0),1),
            "Bajas Año":   int(to_num(gv(rec,"Bajas_año_movil","BajasAñoMovil","Bajas_anio_movil"),0)),
            "Bajas Ene-Mar": int(to_num(gv(rec,"Bajas_ene_mar","Bajas_ene-mar"),0)),
            "Bajas Oct-Dic": int(to_num(gv(rec,"Bajas_oct_dic","Bajas_oct-dic"),0)),
            "B1 Etapa":    b1,
            "B2 Bajas":    b2,
            "B3 Retención":b3,
            "B4 Tendencia":b4,
            "B5 Gestión":  b5,
            "Puntaje":     total,
            "Nivel":       nivel,
            "_orden":      NIVEL_ORDER[nivel],
        })
    df = pd.DataFrame(rows).sort_values("_orden", ascending=False).drop(columns=["_orden"])
    return df

# ── Excel export ──────────────────────────────────────────────────
def generar_excel(df):
    COLORS = {
        "Crítico":  ("C00000","FFFFFF"),
        "Alto":     ("FF8C00","FFFFFF"),
        "Moderado": ("FFD700","000000"),
        "Bajo":     ("70AD47","FFFFFF"),
    }
    wb = Workbook()
    ws = wb.active
    ws.title = "Risk Score"
    ws.freeze_panes = "A3"
    thin = Side(style="thin", color="BFBFBF")
    brd = Border(left=thin,right=thin,top=thin,bottom=thin)

    ws.merge_cells("A1:Q1")
    ws["A1"] = f"Vistage — Grupos en Riesgo  |  {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A1"].font = Font(bold=True,name="Arial",size=12,color="FFFFFF")
    ws["A1"].fill = PatternFill("solid",start_color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height = 28

    headers = list(df.columns)
    widths   = [12,14,22,10,10,10,10,12,12,12,8,8,8,8,8,10,12]
    for i,(h,w) in enumerate(zip(headers,widths),1):
        c = ws.cell(row=2,column=i,value=h)
        c.font = Font(bold=True,color="FFFFFF",name="Arial",size=10)
        c.fill = PatternFill("solid",start_color="1F4E79")
        c.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
        c.border = brd
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[2].height = 40

    for row_idx, row in enumerate(df.itertuples(index=False), 3):
        nivel = row.Nivel
        bg,fg = COLORS.get(nivel,("FFFFFF","000000"))
        for col_idx, val in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.border = brd
            c.alignment = Alignment(horizontal="center",vertical="center")
            if col_idx >= 16:
                c.font = Font(bold=True,name="Arial",size=10,color=fg)
                c.fill = PatternFill("solid",start_color=bg)
            elif col_idx >= 11:
                c.font = Font(name="Arial",size=10)
                c.fill = PatternFill("solid",start_color="F2F2F2")
            else:
                c.font = Font(name="Arial",size=10)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── UI ────────────────────────────────────────────────────────────
def render_badge(nivel):
    cls = f"nivel-{nivel.lower()}"
    return f'<span class="{cls}">{NIVEL_EMOJI[nivel]} {nivel}</span>'

def main():
    # Header
    st.markdown("""
    <div class="header-box">
        <h2 style="margin:0;font-size:24px">📊 Vistage · Grupos en Riesgo</h2>
        <p style="margin:4px 0 0 0;opacity:0.85;font-size:14px">Risk Score en tiempo real desde Zoho Creator</p>
    </div>
    """, unsafe_allow_html=True)

    # Verificar config
    if not all([CLIENT_ID, CLIENT_SECRET, REFRESH_TOKEN]):
        st.error("⚙️ Faltan credenciales. Configurá los Secrets en Streamlit Cloud (ver instrucciones).")
        st.stop()

    # Sidebar filtros
    with st.sidebar:
        st.image("https://upload.wikimedia.org/wikipedia/commons/thumb/8/8e/Vistage_Logo.svg/320px-Vistage_Logo.svg.png", width=160)
        st.markdown("---")
        st.markdown("### 🔍 Filtros")
        filtro_nivel = st.multiselect(
            "Nivel de riesgo",
            ["Crítico","Alto","Moderado","Bajo"],
            default=["Crítico","Alto","Moderado","Bajo"]
        )
        filtro_programa = st.text_input("Programa (filtrar)", "")
        filtro_chair    = st.text_input("Chair (filtrar)", "")
        st.markdown("---")
        actualizar = st.button("🔄 Actualizar datos", use_container_width=True)
        if actualizar:
            st.cache_data.clear()
        st.markdown(f"<small>Última actualización:<br>{datetime.now().strftime('%d/%m/%Y %H:%M')}</small>", unsafe_allow_html=True)

    # Cargar datos
    with st.spinner("Conectando a Zoho Creator..."):
        result = get_access_token()
token, api_domain = result if isinstance(result, tuple) else (result, "https://www.zohoapis.com")
        if not token:
            st.error("❌ No se pudo obtener acceso a Zoho. Verificá las credenciales.")
            st.stop()

        # Descubrir reportes si no están configurados
        reportes_ok = [v for v in REPORTES.values() if v]
        if not reportes_ok:
            st.warning("⚙️ No hay reportes configurados. Descubriendo reportes disponibles...")
            disponibles = discover_reports(token, api_domain)
            if disponibles:
                st.info("📋 Reportes encontrados en tu app:\n\n" + "\n".join(f"- `{r}`" for r in disponibles))
                st.markdown("Copiá los nombres en los **Secrets** de Streamlit Cloud como `REPORTE_GRUPOS`, `REPORTE_MIEMBROS`, etc.")
            else:
                st.error("No se encontraron reportes. Verificá el nombre de la app y las credenciales.")
            st.stop()

        registros = []
        for nombre, reporte in REPORTES.items():
            if reporte:
                recs = fetch_report(reporte, token, api_domain)
                registros.extend(recs)

    if not registros:
        st.warning("No se encontraron registros. Verificá los nombres de los reportes en los Secrets.")
        st.stop()

    df = build_dataframe(registros)

    # Aplicar filtros
    if filtro_nivel:
        df = df[df["Nivel"].isin(filtro_nivel)]
    if filtro_programa:
        df = df[df["Programa"].str.contains(filtro_programa, case=False, na=False)]
    if filtro_chair:
        df = df[df["Chair"].str.contains(filtro_chair, case=False, na=False)]

    # KPIs
    total = len(df)
    criticos  = len(df[df["Nivel"]=="Crítico"])
    altos     = len(df[df["Nivel"]=="Alto"])
    moderados = len(df[df["Nivel"]=="Moderado"])
    bajos     = len(df[df["Nivel"]=="Bajo"])

    c1,c2,c3,c4,c5 = st.columns(5)
    for col, label, val, color in [
        (c1, "Total grupos",  total,     "#1F4E79"),
        (c2, "🔴 Crítico",    criticos,  "#C00000"),
        (c3, "🟠 Alto",       altos,     "#FF8C00"),
        (c4, "🟡 Moderado",   moderados, "#B8860B"),
        (c5, "🟢 Bajo",       bajos,     "#70AD47"),
    ]:
        col.markdown(f"""
        <div class="metric-card">
            <div style="font-size:13px;color:#666;margin-bottom:6px">{label}</div>
            <div style="font-size:36px;font-weight:bold;color:{color}">{val}</div>
        </div>
        """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Tabla principal
    st.markdown("### 📋 Detalle de grupos")

    # Colorear tabla
    def color_nivel(val):
        colores = {
            "Crítico":  "background-color:#C00000;color:white;font-weight:bold",
            "Alto":     "background-color:#FF8C00;color:white;font-weight:bold",
            "Moderado": "background-color:#FFD700;color:black;font-weight:bold",
            "Bajo":     "background-color:#70AD47;color:white;font-weight:bold",
        }
        return colores.get(val, "")

    def color_puntaje(val):
        if val > 55:   return "background-color:#FFCCCC;font-weight:bold"
        elif val > 35: return "background-color:#FFE5CC;font-weight:bold"
        elif val > 15: return "background-color:#FFF9CC;font-weight:bold"
        return "background-color:#E2EFDA;font-weight:bold"

    styled = (
        df.style
        .applymap(color_nivel, subset=["Nivel"])
        .applymap(color_puntaje, subset=["Puntaje"])
        .format({"Retención %": "{:.1f}%", "Puntaje": "{:.0f}"})
    )
    st.dataframe(styled, use_container_width=True, height=500)

    # Descarga Excel
    st.markdown("---")
    col_dl, col_info = st.columns([1,3])
    with col_dl:
        excel_buf = generar_excel(df)
        st.download_button(
            label="⬇️ Descargar Excel",
            data=excel_buf,
            file_name=f"Grupos_Riesgo_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    with col_info:
        st.markdown("""
        <small style="color:#888">
        El Excel incluye los 5 bloques de puntaje con colores por nivel de riesgo.<br>
        Los datos se actualizan automáticamente desde Zoho Creator cada hora.
        </small>
        """, unsafe_allow_html=True)

    # Metodología
    with st.expander("📖 Metodología del Risk Score"):
        st.markdown("""
        | Bloque | Peso | Criterio |
        |--------|------|----------|
        | **B1 · Etapa del grupo** | /20 | Antigüedad 0 años=20 · 1 año=10 · 2+=0 |
        | **B2 · Bajas tempranas** | /35 | 1er/2do año: 1 baja=15 · 2=25 · 3+=35. Grupos maduros: escala reducida |
        | **B3 · Retención** | /20 | ≥90%=0 · 75-89%=5 · 60-74%=10 · 40-59%=15 · <40%=20 |
        | **B4 · Tendencia reciente** | /15 | Bajas ene-mar vs oct-dic: empeora=15 · igual=5 · mejora=0 |
        | **B5 · Gestión PIP** | /10 | Completar manualmente (actualmente en 0) |

        **Niveles:** 🟢 Bajo (0–15) · 🟡 Moderado (16–35) · 🟠 Alto (36–55) · 🔴 Crítico (56–100)
        """)

if __name__ == "__main__":
    main()
